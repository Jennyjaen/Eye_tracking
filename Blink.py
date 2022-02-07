import sys
import math
import os
from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5 import uic
from PyQt5.QtWidgets import *
from PyQt5.QtMultimedia import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import time
import pyautogui as pa
from win32api import *
from datetime import datetime
import tobii_research as tr
import pygame as pg
import schedule as sc
import pandas as pd
from openpyxl import *
from random import *


#https://developer.tobiipro.com/tobii.research/python/reference/1.1.0.23-beta-g9262468f/gaze_data_8py-example.html
#http://developer.tobiipro.com/python/python-step-by-step-guide.html
'''
{'device_time_stamp': 26987408, 'system_time_stamp': 36628255210, 'left_gaze_point_on_display_area': (0.39376187324523926, 0.3558783531188965), 
    'left_gaze_point_in_user_coordinate_system': (-34.740962982177734, 125.78491973876953, 60.326107025146484), 'left_gaze_point_validity': 1, 
    'left_pupil_diameter': 5.1175384521484375, 'left_pupil_validity': 1, 
    'left_gaze_origin_in_user_coordinate_system': (-62.81155014038086, -77.7467269897461, 617.912109375), 
    'left_gaze_origin_in_trackbox_coordinate_system': (0.6468167901039124, 0.7340949177742004, 0.45188695192337036), 'left_gaze_origin_validity': 1, 
    'right_gaze_point_on_display_area': (0.36757388710975647, 0.3632172644138336),
    'right_gaze_point_in_user_coordinate_system': (-44.10836410522461, 124.39735412597656, 59.82101058959961), 'right_gaze_point_validity': 1, 
    'right_pupil_diameter': 5.021240234375, 'right_pupil_validity': 1, 
    'right_gaze_origin_in_user_coordinate_system': (3.122692108154297, -79.176513671875, 617.5389404296875),
    'right_gaze_origin_in_trackbox_coordinate_system': (0.4947779178619385, 0.7384464144706726, 0.44520360231399536), 'right_gaze_origin_validity': 1}
'''
#Global 변수 선언...
music=False
doit=False
goal=0
howlong=100
time_check=0
now=0
g_gaze_data=None
drag_area=[0,500,500] #device time stamp, h, w
click_info=["None", 0, 0]#forward/backward, time , 연속 click 횟수
first=0 
last=0
mid_eye=[500,500]
slider_geo=[0,0]
window_geo=[0,0]
#forward geo, backward geo [0, 1] 은 얘네의 실질적 위치, [2]는 직전에 보고 있었는지 여부로 결정하자.
forward_geo=[0,0,0]
backward_geo=[0,0,0]
pointer_geo=[500,500]
r_data=[0, 0, 0, False, "None",0]
dwell=False
drag=False
total_roop=0
row=2

class Mymp3(QMainWindow):
    recordingInterval_value = 5
    recording_idx = 0
    my_eyetracker = None
    xPos = 0
    yPos = 0

    eyeCursorTrackingPen = QtGui.QPen(QtCore.Qt.red, 3, QtCore.Qt.DashDotLine)
    eyeCursorNonTrackingPen = QtGui.QPen(QtCore.Qt.red, 3, QtCore.Qt.DashDotLine)
    eyeCursorPen = eyeCursorTrackingPen

    isNotTracking= False
    oldPos = QtCore.QPoint()
    painter = QPainter()
    prevNow = datetime.now()

    monitorWidth = 1920
    monitorHeight = 1080

    def __init__(self):
        super().__init__()
        self.initUI()
        self.initEyeTracker()
        #self.ui=uic.loadUi("multi_1.ui", self)
        self.player=QMediaPlayer()
        self.playlist=QMediaPlaylist(self.player)
        self.player.setPlaylist(self.playlist)

        self.player.durationChanged.connect(self.duration_Changed)
        self.player.positionChanged.connect(self.position_Changed)
  
    def initUI(self):
        self.setWindowTitle("Music Player")
        self.setGeometry(300,300,1000,600)
        btn=QPushButton("Add music", self)
        btn.move(20,20)
        btn.clicked.connect(self.btn_clicked)

        self.pointer=QPushButton("", self)
        self.pointer.setGeometry(130,110,40,40)
        self.pointer.setStyleSheet("QPushButton{background-color: green; border-radius : 20; border : 2px solid black;}")
        self.pointer.pressed.connect(self.button_move)

        self.forward=QPushButton("▶", self)
        self.backward=QPushButton("◀", self)
        self.forward.move(700, 110)
        self.forward.resize(80, 80)
        self.backward.move(20, 110)
        self.backward.resize(80, 80)
        self.forward.clicked.connect(self.go_front)
        self.backward.clicked.connect(self.go_back)

        self.where1=QLabel(self)
        self.where1.move(725, 80)
        self.where1.setFont(QFont('Arial', 20))
        self.where2=QLabel(self)
        self.where2.move(50, 80)
        self.where2.setFont(QFont('Arial', 20))

        self.commit=QPushButton("Start",self)
        self.commit.move(150,300)
        self.commit.resize(40,30)
        self.commit.clicked.connect(self.get_goal)
        self.goal=QLabel(self)
        self.goal.setFont(QFont('Arial', 30))
        self.goal.resize(50,50)
        self.goal.move(370,200)

        self.slider_check=QPushButton("", self)
        self.slider_check.setGeometry(130,110,40,40)
        self.slider_check.setGeometry(200,110,40,40)
        self.slider_check.setStyleSheet("QPushButton{background-color:rgb(30,144,255); border-radius : 20; border : 2px solid black;}")
        self.inside_slider=QPushButton("", self)
        self.inside_slider.setGeometry(207,117,26,26)
        self.inside_slider.setStyleSheet("QPushButton{background-color: rgb(135,206,250); border-radius: 13;}")
        self.slider_check.setVisible(False)
        self.inside_slider.setVisible(False)

    def initEyeTracker(self):
        found_eyetrackers = tr.find_all_eyetrackers()

        self.my_eyetracker = found_eyetrackers[0]
        print("Address: " + self.my_eyetracker.address)
        print("Model: " + self.my_eyetracker.model)
        print("Name (It's OK if this is empty): " + self.my_eyetracker.device_name)
        print("Serial number: " + self.my_eyetracker.serial_number)

        self.my_eyetracker.subscribe_to(tr.EYETRACKER_GAZE_DATA, gaze_data_callback, as_dictionary= True)

    #움직이긴 하는데 한번에 움직임
    def paintEvent(self, event):
        global howlong, now, g_gaze_data, total_roop
        qp=QPainter()
        qp.begin(self)
        self.pbar=self.progressbar(qp)
        self.update_geo()
        #pa.moveTo(backward_geo[0]+80, backward_geo[1]+80)
        if(doit==True and total_roop<150):
            self.click_drag()
        elif(total_roop>=150):
            self.goal.setText("Experiment Done.")
        qp.end()

    def button_move(self):
        global howlong, click_area, drag
        #print("click is done either")
        while(GetAsyncKeyState(0x01) & 0x8000):
            mx,_ = pa.position()
            ax=self.geometry().x()
            #ay=self.geometry().y()
            depart=mx-ax-25
            
            #문제: 얘가 async 하게 움직이지 않음.
            #어떤때는 잘 재생되는데, 어떤때는 멈춰있음. 다른 action 취해주면 또 다시 잘됨.
            if(depart>125 and depart<625):
                self.pointer.move(depart, 105)
                self.update_point()
                self.player.setPosition(int((depart-125)*howlong/500))
            elif(depart<=125):
                self.pointer.move(125, 105)
                self.update_point()
                self.player.setPosition(1)
            elif(depart>=625):
                self.pointer.move(625, 105)
                self.update_point()
                self.player.setPosition(howlong)

    #실험 초기화.
    def get_goal(self):
        global goal, doit, music, row, total_roop
        if(not doit):
        #total_roop=0
            print(total_roop)
            total_roop+=1
            self.player.setPosition(0)
            goal = randint(1, 100)
            while((goal%5)==0):
                goal=randint(1, 100)
            self.goal.setText(str(goal))
            doit=True
            
    def update_point(self):
        global pointer_geo
        px=self.pointer.geometry().x()
        py=self.pointer.geometry().y()
        ax=self.geometry().x()
        ay=self.geometry().y()
        #print("updated", px)
        pointer_geo=[px+ax+25, py+ay+25]

    def update_geo(self):
        #print("updated")
        global slider_geo, window_geo, forward_geo, backward_geo
        ax=self.geometry().x()
        ay=self.geometry().y()
        fx=self.forward.geometry().x()
        fy=self.forward.geometry().y()
        bx=self.backward.geometry().x()
        by=self.backward.geometry().y()
        
        slider_geo=[ax+150, ay+130]
        window_geo=[ax, ay]
        forward_geo=[fx+ax,fy+ay,0]
        backward_geo=[bx+ax, by+ay,0]

        #print(forward_geo, backward_geo)

    def progressbar(self, qp):
        global dwell, click_area, window_geo
        pen=QPen(Qt.black,3,Qt.SolidLine)
        brush=QBrush(Qt.VerPattern)
        qp.setPen(pen)
        qp.drawLine(150,130,650,130)
        #눈금.
        pen=QPen(Qt.gray,1, Qt.SolidLine)
        qp.setPen(pen)
        for i in range(100):
            qp.drawLine(150+5*i, 120, 150+i*5, 140)

        pen=QPen(Qt.black, 2, Qt.SolidLine)
        qp.setPen(pen)
        for i in range(21):
            qp.drawLine(150+i*25, 115, 150+i*25, 145)   

        if(dwell==True):
            qp.setBrush(QBrush(Qt.red, Qt.SolidPattern))
            hello_future=click_area[1]-window_geo[0]
            qp.drawEllipse(hello_future, 110, 20, 20)
            self.loudness.move(hello_future-5, 105)
            self.loudness.setText(str(int(hello_future*100/howlong)))

    def duration_Changed(self, msec):
        global howlong
        howlong=msec
        #self.slider.setRange(0,msec)

    def position_Changed(self, position):
        global now, howlong, doit, time_check
        now=position
        self.where1.setText(str(int(now*100/howlong)))
        self.where2.setText(str(int(now*100/howlong)))

        rate=now/howlong
        #print(rate)
        self.pointer.setGeometry(int(130+rate*500),110,40,40)
        self.update_point()
        #self.slider.setValue(position)

    def btn_clicked(self):
        global music
        files=QFileDialog.getOpenFileNames(None, 'Get Audio', filter='Audio Files (*.mp3 *.ogg *.wav)')
        url=QUrl.fromLocalFile(files[0][0])
        if(url!=None):
            music=True
            print(url)
        self.playlist.addMedia(QMediaContent(url))
        self.player.setPlaylist(self.playlist)
        self.playlist.setCurrentIndex(0)
        #self.player.play()

    def go_front(self):
        global now, howlong
        #print("front")
        percent=math.ceil(howlong/100)
        if(howlong-now>percent):
            self.player.setPosition(now+percent)
        else:
            self.player.setPosition(howlong)

    def go_back(self):
        global now, pointer_geo
        #print("backward")
        percent=math.ceil(howlong/100)
        if(now>percent):
            self.player.setPosition(now-percent)
            
        else:
            self.player.setPosition(0)

    #눈이 감겼는지, 감겼다면 몇 초째 감겨있는지 확인.
    def closed_check(self):
        global g_gaze_data, first, last
        closed=False
        
        if(g_gaze_data!=None):
            l_dia=g_gaze_data['left_pupil_diameter']
            l_valid=g_gaze_data['left_pupil_validity']
            r_dia=g_gaze_data['right_pupil_diameter']
            r_val=g_gaze_data['right_pupil_validity']

            if(math.isnan(l_dia) and l_valid==0 and math.isnan(r_dia) and r_val==0):
                closed=True
                if(first == 0):
                    first=time.time()
                last=time.time()

        if(closed==False):
            first = 0
            return False
        else:
            temp= last - first              
            return temp
    
    def click_drag(self):
        #print(time.time())
        global g_gaze_data, window_geo, forward_geo, backward_geo, pointer_geo, slider_geo, click_info, drag_area, now, howlong, goal, time_check, doit
        global r_data, total_roop, row
        
        if(int(now*100/howlong)==goal and time_check==0):
            #self.player.pause()
            time_check=time.time()
        if(time_check!=0  and int(now*100/howlong)==goal):
            #print(time.time()-time_check)
            if(time.time()-time_check>5):
                time_check=0
                #self.player.pause()
                self.player.setPosition(0)
                row+=3
                doit=False #Visualizer 움직임 멈춤
            #total_roop+=1

        if(time_check!=0 and int(now*100/howlong)!=goal):
            time_check=time.time()

        r_data[0]=time.time()
        r_data[3]=self.closed_check()
        r_data[5]=int(now*100/howlong)
        if(g_gaze_data!=None):
            left_eye = g_gaze_data['left_gaze_point_on_display_area']
            right_eye = g_gaze_data['right_gaze_point_on_display_area']
            mid_eye=[((left_eye[0]+right_eye[0])*1920/2), ((left_eye[1]+right_eye[1])*1080/2)]
            r_data[1]=mid_eye[0]-self.geometry().x()
            r_data[2]=mid_eye[1]-self.geometry().y()
            if(math.isnan(mid_eye[0]) or math.isnan(mid_eye[1])):
                left_eye = g_gaze_data['left_gaze_point_on_display_area']
            else:
                pa.moveTo(int(mid_eye[0]), int(mid_eye[1]))
                if(forward_geo[0]<=mid_eye[0] and forward_geo[0]+80>=mid_eye[0] and forward_geo[1]<=mid_eye[1] and forward_geo[1]+80>=mid_eye[1]):
                    self.backward.setStyleSheet("QPushButton{background-color: rgba(220,220,220,0);}")
                    #print("looking forward")
                    if(click_info[0]=="F"):
                        inter=time.time()-click_info[1]
                        print(inter)
                        if(inter>=0.7):
                            self.forward.setStyleSheet("QPushButton{background-color: rgba(30,144,255,255);}")
                            if(inter>=0.7+(0.15+math.floor(total_roop/30)*0.2)*click_info[2]):
                                self.go_front()
                                r_data[4]="Front button fixation"
                                print("click front")
                                click_info[2]+=1
                        else:
                            #print("during")
                            self.grad_button("F", inter)

                    else:
                        click_info=["F", time.time(), 0]
                        r_data[4]="Front button enter"
                        print("forward renewal")
                    
                elif(backward_geo[0]<=mid_eye[0] and backward_geo[0]+80>=mid_eye[0] and backward_geo[1]<=mid_eye[1] and backward_geo[1]+80>=mid_eye[1]): 
                    self.forward.setStyleSheet("QPushButton{background-color: rgba(220,220,220,0); border : 1px solid black;}")
                    if(click_info[0]=="B"):
                        inter=time.time()-click_info[1]
                        if(inter>=0.7):
                            self.backward.setStyleSheet("QPushButton{background-color: rgba(220,220,220,0); border : 1px solid black;}")
                            if(inter>=0.7+(0.15+math.floor(total_roop/30)*0.2)*click_info[2]):
                                self.go_back()
                                print("click back")
                                r_data[4]="Back button fixation"
                                click_info[2]+=1
                        else:
                            #print("during")
                            self.grad_button("B", inter)
                    else:
                        click_info=["B", time.time(), 0]
                        print("backward renewal")
                        r_data[4]="Back button enter"
                    
                elif(slider_geo[1]-10<=mid_eye[0] and mid_eye[0]<=slider_geo[1]+510 and (abs(mid_eye[1]-slider_geo[1])<30)):
                    self.forward.setStyleSheet("QPushButton{background-color: rgba(220,220,220,0); border : 1px solid black;}")
                    self.backward.setStyleSheet("QPushButton{background-color: rgba(220,220,220,0); border : 1px solid black;}")
                    if(drag_area[0]==0):
                        drag_area=[time.time(), mid_eye[0], mid_eye[1]]
                    if(abs(mid_eye[0]-drag_area[1])<60):
                        if(time.time()-drag_area[0]>=0.7):
                            self.slider_check.setVisible(False)
                            self.inside_slider.setVisible(False)
                            #self.slider.setVisible(False)
                            print("drag start")
                            r_data[4]="slider drag"
                            ax=self.geometry().x()
                            depart=drag_area[1]-ax-25
                            
                            if(depart>125 and depart<625):
                                self.pointer.move(depart, 105)
                                self.update_point()
                                self.player.setPosition(int((depart-125)*howlong/500))
                            elif(depart<=125):
                                self.pointer.move(125, 105)
                                self.update_point()
                                self.player.setPosition(1)
                            elif(depart>=625):
                                self.pointer.move(625, 105)
                                self.update_point()
                                self.player.setPosition(howlong)
                        else:
                            r_data[4]="slider enter"
                            ax=self.geometry().x()
                            self.grad_slider(drag_area[1]-ax,time.time()-drag_area[0] )
                    else:
                        drag_area=[time.time(), mid_eye[0], mid_eye[1]]
                        self.slider_check.setVisible(False)
                        self.inside_slider.setVisible(False)

                else:
                    click_info=["None", 0,0]
                    drag_area=[0,500,500]
                    self.forward.setStyleSheet("QPushButton{background-color: rgba(220,220,220,0); border : 1px solid black;}")
                    self.backward.setStyleSheet("QPushButton{background-color: rgba(220,220,220,0); border : 1px solid black;}")
                    #self.slider_check.setVisible(False)
                
        self.record(r_data)

    def grad_button(self, fb, time):
        prog=time/0.7
        stop_1 = str(prog - 0.001)
        stop_2 = str(prog)

        styleSheet="QPushButton{background-color: qconicalgradient(cx:0.5, cy:0.5, angle: 90 stop:{STOP_1} rgba(30,144,255,255), stop:{STOP_2} rgba(220,220,220,0)); border : 1px solid black ;}"
        newStyleSheet=styleSheet.replace("{STOP_1}", stop_1).replace("{STOP_2}", stop_2)
        if(fb=="F"):
            self.forward.setStyleSheet(newStyleSheet)
            print("changing forward")
        elif(fb=="B"):
            self.backward.setStyleSheet(newStyleSheet)
            print("changing backward")

    def grad_slider(self, point, time):
        global howlong
        print("hi")
        prog=time/0.7
        stop_1 = str(prog - 0.001)
        stop_2 = str(prog)
        ax=self.geometry().x()
        self.slider_check.move(point-25,105)
        self.slider_check.setVisible(True)
        self.inside_slider.setVisible(True)
        styleSheet="QPushButton{background-color: qconicalgradient(cx:0.5, cy:0.5, angle: 90 stop:{STOP_1} rgba(30,144,255,255), stop:{STOP_2} rgba(0,0,0,0)); border-radius : 20;}"
        newStyleSheet=styleSheet.replace("{STOP_1}", stop_1).replace("{STOP_2}", stop_2)
        self.slider_check.setStyleSheet(newStyleSheet)
        self.inside_slider.move(point-18,112)
        self.inside_slider.setText(str(int((point-150)*howlong/500)))

    def record(self,data):
        global total_roop, row
        #print(total_roop)
        if(total_roop<30):
            if(total_roop==0):
                row=2
            s15.cell(row, 1).value=data[0]
            s15.cell(row, 2).value=data[1]
            s15.cell(row, 3).value=data[2]
            s15.cell(row, 4).value=data[3]
            s15.cell(row, 5).value=data[4]
            s15.cell(row, 6).value=data[5]

        elif(total_roop>=30 and total_roop<60):
            if(total_roop==30):
                row=2
            s35.cell(row, 1).value=data[0]
            s35.cell(row, 2).value=data[1]
            s35.cell(row, 3).value=data[2]
            s35.cell(row, 4).value=data[3]
            s35.cell(row, 5).value=data[4]
            s35.cell(row, 6).value=data[5]

        elif(total_roop>=60 and total_roop<90):
            if(total_roop==60):
                row=2
            s55.cell(row, 1).value=data[0]
            s55.cell(row, 2).value=data[1]
            s55.cell(row, 3).value=data[2]
            s55.cell(row, 4).value=data[3]
            s55.cell(row, 5).value=data[4]
            s55.cell(row, 6).value=data[5]

        elif(total_roop>=90 and total_roop<120):
            if(total_roop==90):
                row=2
            s75.cell(row, 1).value=data[0]
            s75.cell(row, 2).value=data[1]
            s75.cell(row, 3).value=data[2]
            s75.cell(row, 4).value=data[3]
            s75.cell(row, 5).value=data[4]
            s75.cell(row, 6).value=data[5]

        elif(total_roop>=120 and total_roop<150):
            if(total_roop==120):
                row=2
            s95.cell(row, 1).value=data[0]
            s95.cell(row, 2).value=data[1]
            s95.cell(row, 3).value=data[2]
            s95.cell(row, 4).value=data[3]
            s95.cell(row, 5).value=data[4]
            s95.cell(row, 6).value=data[5]

        else:
            raise NameError(['invalid interval'])
        
        row+=1
        #print("hi")

    def closeEvent(self, event):
        self.my_eyetracker.unsubscribe_from(tr.EYETRACKER_GAZE_DATA, gaze_data_callback)
        wb.save('test.xlsx')
        print('destroy Visualizer')

    def keyPressEvent(self, event):
        key = event.key()
        if key == Qt.Key_Space:
            print('space')

def gaze_data_callback(gaze_data):
    # Print gaze points of left and right eye
    global g_gaze_data, mid_eye
    g_gaze_data=gaze_data

    left_eye = gaze_data['left_gaze_point_on_display_area']
    right_eye = gaze_data['right_gaze_point_on_display_area']
    mid_eye=[((left_eye[0]+right_eye[0])*1920/2), ((left_eye[1]+right_eye[1])*1080/2)]
    

#main에서 설정한 timer로 0.03초에 한번씩 timer Event 발생. 
#timer event에서 repaint 하기 때문에 myWindow내의 paintevent를 0.03초마다 발생시킬 수 있다!
def timerEvent():
    myWindow.repaint()

def exit():
    print('exit')
    timer.stop()
    ex.destroy()
    app.exit(0)


if __name__=="__main__":
    #global doit
    wb=Workbook()
    s15=wb.active
    s15.title="0.15"
    s35=wb.create_sheet("0.35")
    s55=wb.create_sheet("0.55")
    s75=wb.create_sheet("0.75")
    s95=wb.create_sheet("0.95")

    app=QApplication(sys.argv)
    myWindow=Mymp3()
    
    timer = QtCore.QTimer()
    timer.timeout.connect(timerEvent)
    timer.start(0.03)

    myWindow.show()
    sys.exit(app.exec_())
    exit()